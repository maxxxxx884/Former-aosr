import os
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from path_manager import PathManager


def is_filled(value):
    """Проверяет, что значение не является пустым"""
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


class KS2Processor:
    """Класс для обработки вставки проектной сметы в шаблон КС-2"""

    def __init__(self, template_path, source_path, output_path):
        self.template_path = template_path
        self.source_path = source_path
        self.output_path = output_path

    def find_ks2_sheet(self, workbook):
        """Находит лист, начинающийся с 'КС-2'"""
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("КС-2"):
                return workbook[sheet_name]
        raise ValueError("Не найден лист, начинающийся с 'КС-2'")

    def get_table_dimensions(self, source_sheet):
        """Определяет размеры таблицы из исходного файла"""
        max_row = 0
        max_col = 0

        for row in source_sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)

        return max_row, max_col

    def shift_rows(self, sheet, start_row, rows_to_insert):
        """Сдвигает строки вниз начиная с start_row"""
        print(f"  Сдвиг строк с {start_row} на {rows_to_insert} позиций вниз...")

        # Получаем максимальную используемую строку
        max_row = sheet.max_row

        # Сдвигаем строки снизу вверх, чтобы избежать перезаписи
        for row_idx in range(max_row, start_row - 1, -1):
            for col_idx in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=row_idx, column=col_idx)
                target_cell = sheet.cell(row=row_idx + rows_to_insert, column=col_idx)

                # Копируем значение
                target_cell.value = source_cell.value

                # Копируем форматирование
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()

        # Очищаем освободившиеся строки
        for row_idx in range(start_row, start_row + rows_to_insert):
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=row_idx, column=col_idx).value = None

    def shift_range_left(self, sheet, range_start, range_end, columns_to_shift):
        """
        Сдвигает диапазон ячеек влево на указанное количество столбцов
        range_start, range_end: кортежи (строка, столбец)
        """
        print(f"  Сдвиг диапазона {get_column_letter(range_start[1])}{range_start[0]}:" +
              f"{get_column_letter(range_end[1])}{range_end[0]} влево на {columns_to_shift} столбцов...")

        for row_idx in range(range_start[0], range_end[0] + 1):
            for col_idx in range(range_start[1], range_end[1] + 1):
                source_cell = sheet.cell(row=row_idx, column=col_idx)
                target_col = col_idx - columns_to_shift

                if target_col >= 1:  # Проверяем, что не выходим за границы
                    target_cell = sheet.cell(row=row_idx, column=target_col)

                    # Копируем значение
                    target_cell.value = source_cell.value

                    # Копируем форматирование
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()

                # Очищаем исходную ячейку
                source_cell.value = None

    def insert_table(self, target_sheet, source_sheet, start_row=20):
        """Вставляет таблицу из source_sheet в target_sheet начиная со start_row"""
        print(f"  Вставка таблицы начиная со строки {start_row}...")

        # Получаем размеры исходной таблицы
        source_rows, source_cols = self.get_table_dimensions(source_sheet)
        print(f"  Размеры вставляемой таблицы: {source_rows} строк × {source_cols} столбцов")

        # Копируем данные
        for row_idx in range(1, source_rows + 1):
            for col_idx in range(1, source_cols + 1):
                source_cell = source_sheet.cell(row=row_idx, column=col_idx)
                target_cell = target_sheet.cell(row=start_row + row_idx - 1, column=col_idx)

                # Копируем значение
                target_cell.value = source_cell.value

                # Копируем форматирование
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()

        return source_rows, source_cols

    def process(self):
        """Основной метод обработки"""
        print("🚀 === НАЧАЛО ОБРАБОТКИ ===\n")
        print(f"📄 Шаблон: {os.path.basename(self.template_path)}")
        print(f"📊 Исходные данные: {os.path.basename(self.source_path)}")

        try:
            # Загружаем файлы
            print("\n📥 Загрузка файлов...")
            template_wb = load_workbook(self.template_path)
            source_wb = load_workbook(self.source_path)

            # Находим нужные листы
            ks2_sheet = self.find_ks2_sheet(template_wb)
            source_sheet = source_wb.active

            print(f"✅ Найден лист шаблона: '{ks2_sheet.title}'")
            print(f"✅ Используется исходный лист: '{source_sheet.title}'")

            # Получаем размеры вставляемой таблицы
            source_rows, source_cols = self.get_table_dimensions(source_sheet)

            # 1. Сдвигаем строки в шаблоне
            print(f"\n🔄 Сдвиг строк в шаблоне...")
            self.shift_rows(ks2_sheet, start_row=20, rows_to_insert=source_rows)

            # 2. Вставляем таблицу
            print(f"\n📋 Вставка данных...")
            inserted_rows, inserted_cols = self.insert_table(ks2_sheet, source_sheet, start_row=20)

            # 3. Проверяем, нужно ли сдвигать области G1:H18 и E12:F18
            # Столбец H это 8-й столбец
            if inserted_cols > 8:
                columns_to_shift = inserted_cols - 8
                print(f"\n⬅️  Вставленная таблица выходит за столбец H")
                print(f"  Необходимо сдвинуть области влево на {columns_to_shift} столбцов")

                # Сдвигаем область G1:H18 (столбцы 7-8, строки 1-18)
                self.shift_range_left(ks2_sheet,
                                      range_start=(1, 7),
                                      range_end=(18, 8),
                                      columns_to_shift=columns_to_shift)

                # Сдвигаем область E12:F18 (столбцы 5-6, строки 12-18)
                self.shift_range_left(ks2_sheet,
                                      range_start=(12, 5),
                                      range_end=(18, 6),
                                      columns_to_shift=columns_to_shift)
            else:
                print(f"\n✅ Вставленная таблица заканчивается на столбце {get_column_letter(inserted_cols)}")
                print(f"  Сдвиг областей G1:H18 и E12:F18 не требуется")

            # Сохраняем результат
            print(f"\n💾 Сохранение результата...")
            template_wb.save(self.output_path)

            print(f"\n🎉 === ОБРАБОТКА ЗАВЕРШЕНА ===")
            print(f"✅ Результат сохранен: {os.path.basename(self.output_path)}")

            return True

        except Exception as e:
            print(f"\n❌ Ошибка: {str(e)}")
            raise


class ToolTip:
    """Класс для создания всплывающих подсказок"""

    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)
        self.tooltip = None

    def on_enter(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 20
        y += self.widget.winfo_rooty() + 20

        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tooltip, text=self.text,
                         background="lightyellow",
                         relief="solid", borderwidth=1,
                         font=("Arial", "8", "normal"))
        label.pack()

    def on_leave(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None


class KS2Application:
    """Главное приложение с GUI"""

    def __init__(self):
        self.root = TkinterDnD.Tk()
        self.root.title("📊 Вставка проектной сметы в КС-2")
        self.root.geometry("800x800")

        # Инициализация менеджера путей
        self.path_manager = PathManager()
        saved_paths = self.path_manager.load_paths()

        # Переменные для путей
        self.template_path = tk.StringVar(value=saved_paths.get("ks2_template", ""))
        self.source_path = tk.StringVar(value=saved_paths.get("source_file", ""))
        self.output_path = tk.StringVar(value=saved_paths.get("output_file", ""))

        # Переменные валидации
        self.validation_vars = {
            'template': tk.BooleanVar(),
            'source': tk.BooleanVar(),
            'output': tk.BooleanVar()
        }

        self.setup_ui()

        # Валидация после загрузки
        self.root.after(100, lambda: [
            self.validate_path(self.template_path, self.validation_vars['template'], True),
            self.validate_path(self.source_path, self.validation_vars['source'], True),
            self.validate_path(self.output_path, self.validation_vars['output'], False)
        ])

    def setup_ui(self):
        """Создание интерфейса"""
        # Настройка стилей
        style = ttk.Style()
        style.theme_use('clam')

        # Заголовок
        header_frame = ttk.Frame(self.root)
        header_frame.pack(fill=tk.X, padx=20, pady=(20, 10))

        title_label = ttk.Label(header_frame,
                                text="📊 Вставка проектной сметы в шаблон КС-2",
                                font=("Arial", 16, "bold"))
        title_label.pack()

        subtitle_label = ttk.Label(header_frame,
                                   text="Автоматическая вставка данных с учетом форматирования",
                                   font=("Arial", 10))
        subtitle_label.pack()

        # Основная область
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Секции выбора файлов
        self.create_file_section(main_frame,
                                 "📄 Шаблон КС-2",
                                 "Файл Excel с шаблоном КС-2 (лист должен начинаться с 'КС-2')",
                                 self.template_path,
                                 self.validation_vars['template'])

        self.create_file_section(main_frame,
                                 "📊 Проектная смета",
                                 "Файл Excel с проектной сметой для вставки",
                                 self.source_path,
                                 self.validation_vars['source'])

        self.create_file_section(main_frame,
                                 "💾 Файл результата",
                                 "Путь для сохранения обработанного файла",
                                 self.output_path,
                                 self.validation_vars['output'],
                                 is_output=True)

        # Информационная панель
        info_frame = ttk.LabelFrame(main_frame, text="ℹ️ Информация", padding=(15, 10))
        info_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))

        info_text = tk.Text(info_frame, height=8, wrap=tk.WORD, font=("Arial", 9))
        info_text.pack(fill=tk.BOTH, expand=True)

        info_content = """Программа выполняет следующие действия:

1. 📥 Загружает шаблон КС-2 и файл с проектной сметой
2. 🔍 Находит лист, начинающийся с 'КС-2' в шаблоне
3. 📏 Определяет размеры вставляемой таблицы
4. ⬇️  Сдвигает строки в шаблоне начиная с 20-й на высоту таблицы
5. 📋 Вставляет данные из проектной сметы с сохранением форматирования
6. ⬅️  При необходимости сдвигает области G1:H18 и E12:F18 влево
7. 💾 Сохраняет результат в указанный файл

💡 Важно: Все форматирование (шрифты, границы, заливка) сохраняется!"""

        info_text.insert(tk.END, info_content)
        info_text.config(state=tk.DISABLED)

        # Нижняя панель
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(fill=tk.X, padx=20, pady=(0, 20))

        # Статус
        self.status_label = ttk.Label(bottom_frame,
                                      text="📋 Заполните все поля",
                                      font=("Arial", 10))
        self.status_label.pack(side=tk.LEFT)

        # Кнопки
        button_frame = ttk.Frame(bottom_frame)
        button_frame.pack(side=tk.RIGHT)

        self.process_btn = ttk.Button(button_frame,
                                      text="🚀 Обработать",
                                      command=self.process_files,
                                      state='disabled')
        self.process_btn.pack(side=tk.LEFT, padx=(10, 5))

        exit_btn = ttk.Button(button_frame,
                              text="❌ Выход",
                              command=self.root.quit)
        exit_btn.pack(side=tk.LEFT, padx=5)

        # Горячие клавиши
        self.root.bind('<Return>', lambda e: self.process_files() if self.process_btn['state'] == 'normal' else None)
        self.root.bind('<Escape>', lambda e: self.root.quit())

    def create_file_section(self, parent, title, description, variable, validation_var, is_output=False):
        """Создает секцию для выбора файла"""
        section_frame = ttk.LabelFrame(parent, text=title, padding=(10, 5))
        section_frame.pack(fill=tk.X, pady=8)

        # Описание
        desc_label = ttk.Label(section_frame, text=description,
                               font=("Arial", 9), foreground="gray")
        desc_label.pack(anchor="w")

        # Поле ввода
        input_frame = ttk.Frame(section_frame)
        input_frame.pack(fill=tk.X, pady=(5, 0))

        entry = ttk.Entry(input_frame, textvariable=variable, font=("Arial", 9))
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Индикатор валидации
        status_label = ttk.Label(input_frame, text="❌", font=("Arial", 12))
        status_label.pack(side=tk.LEFT, padx=(5, 0))

        # Кнопка обзора
        def browse():
            if is_output:
                path = filedialog.asksaveasfilename(
                    parent=self.root,
                    title=title,
                    defaultextension=".xlsx",
                    filetypes=[('Excel файлы', '*.xlsx'), ('Все файлы', '*.*')]
                )
            else:
                path = filedialog.askopenfilename(
                    parent=self.root,
                    title=title,
                    filetypes=[('Excel файлы', '*.xlsx *.xls'), ('Все файлы', '*.*')]
                )

            if is_filled(path):
                variable.set(path)
                self.validate_path(variable, validation_var, not is_output)

        browse_btn = ttk.Button(input_frame, text="📁 Обзор", command=browse)
        browse_btn.pack(side=tk.LEFT, padx=(5, 0))

        # Drag & Drop
        entry.drop_target_register(DND_FILES)
        entry.dnd_bind('<<Drop>>', lambda e: self.on_drop(e, variable, validation_var, not is_output))

        # Валидация при изменении
        variable.trace('w', lambda *args: self.validate_path(variable, validation_var, not is_output))

        # Обновление индикатора
        def update_status(*args):
            if validation_var.get():
                status_label.config(text="✅", foreground="green")
            else:
                status_label.config(text="❌", foreground="red")

        validation_var.trace('w', update_status)

    def on_drop(self, event, var, validation_var, must_exist):
        """Обработка Drag & Drop"""
        file_path = event.data.strip('{}')
        if is_filled(file_path):
            var.set(file_path)
            self.validate_path(var, validation_var, must_exist)

    def validate_path(self, var, validation_var, must_exist=True):
        """Валидация пути"""
        path = var.get()

        if not is_filled(path):
            validation_var.set(False)
            self.update_process_button()
            return

        if must_exist:
            validation_var.set(os.path.exists(path) and os.path.isfile(path))
        else:
            # Для выходного файла проверяем только, что путь валидный
            validation_var.set(len(path) > 0)

        self.update_process_button()

    def update_process_button(self):
        """Обновление состояния кнопки обработки"""
        all_valid = all(var.get() for var in self.validation_vars.values())

        if all_valid:
            self.process_btn.config(state='normal')
            self.status_label.config(text="✅ Готово к обработке", foreground="green")
        else:
            self.process_btn.config(state='disabled')
            self.status_label.config(text="📋 Заполните все поля", foreground="orange")

    def process_files(self):
        """Запуск обработки файлов"""
        # Сохраняем пути
        paths_to_save = {
            "ks2_template": self.template_path.get().strip(),
            "source_file": self.source_path.get().strip(),
            "output_file": self.output_path.get().strip()
        }
        self.path_manager.save_paths(paths_to_save)

        # Создаем окно прогресса
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Обработка...")
        progress_window.geometry("500x200")
        progress_window.transient(self.root)
        progress_window.grab_set()

        # Центрируем окно
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
        y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
        progress_window.geometry(f"+{x}+{y}")

        # Текстовое поле для вывода
        text_frame = ttk.Frame(progress_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 9))
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.config(yscrollcommand=scrollbar.set)

        # Перенаправляем print в текстовое поле
        import sys
        from io import StringIO

        old_stdout = sys.stdout
        sys.stdout = StringIO()

        def update_text():
            output = sys.stdout.getvalue()
            text_widget.delete(1.0, tk.END)
            text_widget.insert(tk.END, output)
            text_widget.see(tk.END)
            progress_window.update()

        try:
            # Создаем процессор
            processor = KS2Processor(
                self.template_path.get().strip(),
                self.source_path.get().strip(),
                self.output_path.get().strip()
            )

            # Обновляем GUI во время обработки
            def process_with_updates():
                try:
                    processor.process()
                    update_text()

                    messagebox.showinfo(
                        "Успех",
                        "Обработка успешно завершена!\n\n" +
                        f"Результат сохранен в:\n{self.output_path.get()}",
                        parent=progress_window
                    )
                except Exception as e:
                    update_text()
                    messagebox.showerror(
                        "Ошибка",
                        f"Произошла ошибка при обработке:\n\n{str(e)}",
                        parent=progress_window
                    )
                finally:
                    sys.stdout = old_stdout
                    progress_window.destroy()

            # Запускаем обработку после отображения окна
            progress_window.after(100, process_with_updates)

        except Exception as e:
            sys.stdout = old_stdout
            progress_window.destroy()
            messagebox.showerror(
                "Ошибка",
                f"Произошла ошибка:\n\n{str(e)}",
                parent=self.root
            )

    def run(self):
        """Запуск приложения"""
        # Центрируем главное окно
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (self.root.winfo_width() // 2)
        y = (self.root.winfo_screenheight() // 2) - (self.root.winfo_height() // 2)
        self.root.geometry(f"+{x}+{y}")

        self.root.mainloop()


if __name__ == "__main__":
    app = KS2Application()
    app.run()
